"""从 collected-data.xlsx 构建仓库清单 manifest.json。"""
from __future__ import annotations

import json
import re
import shutil
from collections import Counter
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse

import openpyxl

from ..config import settings
from ..logging import logger
from ..schemas import (
    DeleteRepoResult,
    ImportReport,
    ImportRowResult,
    Manifest,
    ManualRepoInput,
    RepoEntry,
    RepoStatus,
)

# Excel 表头映射（与 collected-data.xlsx 对齐）
COL_YEAR = 0
COL_CONTEST = 1
COL_TRACK = 2
COL_SCHOOL = 3
COL_TEAM = 4
COL_URL = 5


# 保留中英文 / 数字 / 下划线 / 连字符；其他字符（空格、标点、emoji）替换为 -
_SLUG_PATTERN = re.compile(r"[^\w\-\u4e00-\u9fff]+", flags=re.UNICODE)


def _slug(text: str, max_len: int = 30) -> str:
    text = _SLUG_PATTERN.sub("-", text.strip()).strip("-")
    return text[:max_len] or "unknown"


def _detect_host(url: str) -> str:
    try:
        host = urlparse(url).netloc.lower()
    except Exception:
        return "other"
    if "eduxiji" in host:
        return "gitlab.eduxiji.net"
    if "github" in host:
        return "github.com"
    if "gitee" in host:
        return "gitee.com"
    return "other"


def _make_repo_id(year: int, idx: int, team: str) -> str:
    return f"{year}_{idx:03d}_{_slug(team)}"


def load_from_xlsx(xlsx_path: Path) -> list[RepoEntry]:
    """读取 Excel，返回原始 RepoEntry 列表（status = PENDING）。"""
    logger.info(f"读取数据集: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    entries: list[RepoEntry] = []
    per_year_idx: Counter[int] = Counter()

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # 表头
        if not row or row[COL_URL] is None:
            continue
        year = int(row[COL_YEAR])
        contest = str(row[COL_CONTEST] or "").strip()
        track = str(row[COL_TRACK] or "").strip()
        school = str(row[COL_SCHOOL] or "").strip()
        team = str(row[COL_TEAM] or "").strip()
        url = str(row[COL_URL]).strip()

        per_year_idx[year] += 1
        repo_id = _make_repo_id(year, per_year_idx[year], team)

        entries.append(
            RepoEntry(
                year=year,
                contest=contest,
                track=track,
                school=school,
                team=team,
                repo_url=url,
                repo_id=repo_id,
                host=_detect_host(url),  # type: ignore[arg-type]
                status=RepoStatus.PENDING,
            )
        )

    wb.close()
    logger.info(f"共解析 {len(entries)} 条仓库记录")
    return entries


def build_manifest(xlsx_path: Path | None = None, out_path: Path | None = None) -> Manifest:
    """构建并落盘 manifest.json。若已存在则保留 status / local_path 等运行时字段。"""
    settings.ensure_dirs()
    xlsx_path = xlsx_path or settings.dataset_xlsx
    out_path = out_path or settings.manifest_path

    new_entries = load_from_xlsx(xlsx_path)

    # 合并旧 manifest 的运行时状态
    if out_path.exists():
        try:
            old = Manifest.model_validate_json(out_path.read_text(encoding="utf-8"))
            old_map = {e.repo_id: e for e in old.repos}
            merged = 0
            for e in new_entries:
                if e.repo_id in old_map:
                    o = old_map[e.repo_id]
                    e.status = o.status
                    e.local_path = o.local_path
                    e.default_branch = o.default_branch
                    e.head_commit = o.head_commit
                    e.size_bytes = o.size_bytes
                    e.file_count = o.file_count
                    e.cloned_at = o.cloned_at
                    e.error_msg = o.error_msg
                    merged += 1
            logger.info(f"已从旧 manifest 合并 {merged} 条运行时状态")
        except Exception as exc:
            logger.warning(f"旧 manifest 解析失败，将覆盖: {exc}")

    manifest = Manifest(
        source_xlsx=str(xlsx_path),
        total=len(new_entries),
        repos=new_entries,
    )

    out_path.write_text(
        manifest.model_dump_json(indent=2, exclude_none=False),
        encoding="utf-8",
    )
    logger.info(f"manifest 已写入: {out_path}  (total={manifest.total})")
    return manifest


def load_manifest(path: Path | None = None) -> Manifest:
    path = path or settings.manifest_path
    if not path.exists():
        raise FileNotFoundError(f"manifest 不存在: {path}，请先 `osagent manifest build`")
    return Manifest.model_validate_json(path.read_text(encoding="utf-8"))


def save_manifest(manifest: Manifest, path: Path | None = None) -> None:
    path = path or settings.manifest_path
    path.write_text(manifest.model_dump_json(indent=2, exclude_none=False), encoding="utf-8")


def manifest_stats(manifest: Manifest) -> dict:
    by_year = Counter(e.year for e in manifest.repos)
    by_host = Counter(e.host for e in manifest.repos)
    by_status = Counter(e.status.value for e in manifest.repos)
    by_track = Counter(e.track for e in manifest.repos)
    schools = {e.school for e in manifest.repos}
    return {
        "total": manifest.total,
        "by_year": dict(sorted(by_year.items())),
        "by_host": dict(by_host),
        "by_status": dict(by_status),
        "by_track": dict(by_track),
        "schools_count": len(schools),
    }


# ============================================================
#  增量导入 / 单仓添加 / 删除（v0.7）
# ============================================================

def _backup_dir() -> Path:
    p = settings.data_dir / "backups"
    p.mkdir(parents=True, exist_ok=True)
    return p


def backup_manifest(reason: str = "auto") -> Path | None:
    """把当前 manifest.json 复制到 data/backups/manifest-<ts>-<reason>.json。

    没有 manifest 文件就返回 None（首次导入场景）。
    """
    src = settings.manifest_path
    if not src.exists():
        return None
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    dst = _backup_dir() / f"manifest-{ts}-{reason}.json"
    shutil.copy2(src, dst)
    logger.info(f"manifest 已备份到: {dst}")
    return dst


def _max_year_idx(manifest: Manifest) -> dict[int, int]:
    """从现有 manifest 推算每个年份当前最大序号（用于增量续号）。

    repo_id 格式：``<year>_<idx:03d>_<team_slug>``，解析失败的条目跳过。
    """
    max_idx: dict[int, int] = {}
    for r in manifest.repos:
        try:
            parts = r.repo_id.split("_", 2)
            y = int(parts[0])
            i = int(parts[1])
        except (ValueError, IndexError):
            continue
        if i > max_idx.get(y, 0):
            max_idx[y] = i
    return max_idx


def _read_xlsx_rows(xlsx_path: Path) -> list[tuple[int, tuple]]:
    """读 xlsx 返回 [(excel_row_number, row_tuple), ...]，已跳过表头。"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows: list[tuple[int, tuple]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        rows.append((i + 1, row))  # 1-based 行号，方便用户对照 Excel
    wb.close()
    return rows


def import_xlsx_incremental(
    xlsx_path: Path,
    *,
    dry_run: bool = False,
    source_label: str | None = None,
) -> ImportReport:
    """增量导入 xlsx 到 manifest（merge 模式：URL 已存在则跳过）。

    - 不会清空现有仓库
    - 不会改变已存在仓库的 repo_id（即使 team/school 在 Excel 里改名）
    - 新仓的序号从该年份当前最大值 +1 续起
    - dry_run=True 时只返回 ImportReport，不动盘 / 不备份
    - 实际执行时自动备份当前 manifest 到 data/backups/
    """
    settings.ensure_dirs()

    # 1) 加载现有 manifest（不存在则视为空，相当于首次导入）
    if settings.manifest_path.exists():
        manifest = load_manifest()
    else:
        manifest = Manifest(
            source_xlsx=str(xlsx_path),
            total=0,
            repos=[],
        )

    existing_urls: dict[str, str] = {r.repo_url: r.repo_id for r in manifest.repos}
    year_idx = _max_year_idx(manifest)

    # 2) 解析 xlsx
    rows = _read_xlsx_rows(xlsx_path)
    report = ImportReport(
        source=source_label or str(xlsx_path),
        mode="merge",
        dry_run=dry_run,
        total_rows=len(rows),
    )

    new_entries: list[RepoEntry] = []
    seen_in_xlsx: set[str] = set()

    for excel_row, row in rows:
        # ---- 行级解析 + 校验 ----
        if not row or len(row) <= COL_URL or row[COL_URL] is None:
            report.rows.append(ImportRowResult(
                row=excel_row, action="skipped", reason="empty repo_url",
            ))
            report.skipped += 1
            continue

        try:
            year = int(row[COL_YEAR])
        except (TypeError, ValueError):
            report.rows.append(ImportRowResult(
                row=excel_row, action="error",
                reason=f"year 不是数字: {row[COL_YEAR]!r}",
            ))
            report.errors += 1
            continue

        url = str(row[COL_URL]).strip()
        if not url:
            report.rows.append(ImportRowResult(
                row=excel_row, action="skipped", reason="repo_url 为空字符串",
            ))
            report.skipped += 1
            continue

        # 已在 manifest 中：跳过
        if url in existing_urls:
            report.rows.append(ImportRowResult(
                row=excel_row, action="skipped", repo_url=url,
                repo_id=existing_urls[url],
                reason="repo_url 已存在",
            ))
            report.skipped += 1
            continue

        # 本次 xlsx 内重复：跳过后者
        if url in seen_in_xlsx:
            report.rows.append(ImportRowResult(
                row=excel_row, action="skipped", repo_url=url,
                reason="xlsx 内重复（保留第一次出现）",
            ))
            report.skipped += 1
            continue

        contest = str(row[COL_CONTEST] or "").strip() or "操作系统赛"
        track = str(row[COL_TRACK] or "").strip() or "内核实现赛道"
        school = str(row[COL_SCHOOL] or "").strip() or "未知"
        team = str(row[COL_TEAM] or "").strip() or "未知队伍"

        year_idx[year] = year_idx.get(year, 0) + 1
        repo_id = _make_repo_id(year, year_idx[year], team)

        entry = RepoEntry(
            year=year,
            contest=contest,
            track=track,
            school=school,
            team=team,
            repo_url=url,
            repo_id=repo_id,
            host=_detect_host(url),  # type: ignore[arg-type]
            status=RepoStatus.PENDING,
        )
        new_entries.append(entry)
        seen_in_xlsx.add(url)
        report.rows.append(ImportRowResult(
            row=excel_row, action="added",
            repo_id=repo_id, repo_url=url,
        ))
        report.added += 1

    # 3) 实际落盘（dry_run 跳过）
    if not dry_run and new_entries:
        report.backup_path = str(backup_manifest(reason="import-xlsx") or "")
        manifest.repos.extend(new_entries)
        manifest.total = len(manifest.repos)
        manifest.generated_at = datetime.now()
        # source_xlsx 保留旧值；额外记录最近一次导入
        save_manifest(manifest)
        logger.info(
            f"增量导入完成: +{report.added} 条; 跳过 {report.skipped}; "
            f"错误 {report.errors}; manifest.total={manifest.total}"
        )
    elif not dry_run and not new_entries:
        # 没东西要写，仍然写一条备份记录便于回滚（可选：可省略）
        logger.info(
            f"增量导入：无新增 (added=0, skipped={report.skipped}, errors={report.errors})"
        )
    else:
        logger.info(
            f"[dry-run] 预览: +{report.added} 跳过 {report.skipped} 错误 {report.errors}"
        )

    return report


def add_repo_manual(input_data: ManualRepoInput) -> RepoEntry:
    """手工添加一条仓库。URL 已存在则抛 ValueError。"""
    settings.ensure_dirs()

    if settings.manifest_path.exists():
        manifest = load_manifest()
    else:
        manifest = Manifest(
            source_xlsx=str(settings.dataset_xlsx),
            total=0,
            repos=[],
        )

    url = input_data.repo_url.strip()
    if not url:
        raise ValueError("repo_url 不能为空")

    for r in manifest.repos:
        if r.repo_url == url:
            raise ValueError(f"该 URL 已存在: {r.repo_id}")

    year_idx = _max_year_idx(manifest)
    next_idx = year_idx.get(input_data.year, 0) + 1
    repo_id = _make_repo_id(input_data.year, next_idx, input_data.team)

    entry = RepoEntry(
        year=input_data.year,
        contest=input_data.contest.strip() or "操作系统赛",
        track=input_data.track.strip() or "内核实现赛道",
        school=input_data.school.strip() or "未知",
        team=input_data.team.strip() or "未知队伍",
        repo_url=url,
        repo_id=repo_id,
        host=_detect_host(url),  # type: ignore[arg-type]
        status=RepoStatus.PENDING,
    )

    backup_manifest(reason="add-repo")
    manifest.repos.append(entry)
    manifest.total = len(manifest.repos)
    manifest.generated_at = datetime.now()
    save_manifest(manifest)
    logger.info(f"已添加仓库: {repo_id} ({input_data.team})")
    return entry


def delete_repo(repo_id: str, *, purge_data: bool = False) -> DeleteRepoResult:
    """从 manifest 删除一条记录。

    purge_data=True 时连同 ``data/repos/<repo_id>`` 目录、
    ``data/facts/<repo_id>.json``、
    ``data/reports/<repo_id>.{md,html}`` 一起清理；
    compare 报告**不清**（避免误删跨多个 repo 的产物）。
    """
    manifest = load_manifest()
    target = next((r for r in manifest.repos if r.repo_id == repo_id), None)
    if target is None:
        raise ValueError(f"repo_id 未找到: {repo_id}")

    purged: list[str] = []
    skipped: list[str] = []

    # 先备份
    backup_manifest(reason=f"delete-{repo_id}")

    if purge_data:
        # 1) data/repos/<repo_id>/
        repo_dir = settings.repos_dir / repo_id
        if repo_dir.exists():
            try:
                shutil.rmtree(repo_dir)
                purged.append(str(repo_dir))
            except OSError as e:
                skipped.append(f"{repo_dir}: {e}")
        else:
            skipped.append(f"{repo_dir} (不存在)")

        # 2) data/facts/<repo_id>.json
        facts_file = settings.facts_dir / f"{repo_id}.json"
        if facts_file.exists():
            try:
                facts_file.unlink()
                purged.append(str(facts_file))
            except OSError as e:
                skipped.append(f"{facts_file}: {e}")

        # 3) data/reports/<repo_id>.md / .html
        for ext in (".md", ".html"):
            f = settings.reports_dir / f"{repo_id}{ext}"
            if f.exists():
                try:
                    f.unlink()
                    purged.append(str(f))
                except OSError as e:
                    skipped.append(f"{f}: {e}")

    # 从 manifest 移除
    manifest.repos = [r for r in manifest.repos if r.repo_id != repo_id]
    manifest.total = len(manifest.repos)
    manifest.generated_at = datetime.now()
    save_manifest(manifest)
    logger.info(
        f"已删除仓库: {repo_id} (purge_data={purge_data}, "
        f"purged={len(purged)}, skipped={len(skipped)})"
    )
    return DeleteRepoResult(
        repo_id=repo_id,
        deleted=True,
        purged_paths=purged,
        skipped_paths=skipped,
    )
